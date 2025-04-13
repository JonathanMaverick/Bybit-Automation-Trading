from pybit.unified_trading import HTTP
from config import API_KEY, API_SECRET, ACCOUNT_TYPE, MAX_POSITIONS, TIME_THRESHOLD
import pandas as pd
from time import sleep
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from services.discord_service import send_discord

class Bybit:
    def __init__(self):
        self.client = HTTP(
            testnet=False,
            api_key=API_KEY,
            api_secret=API_SECRET,
        )

    def get_balance(self):
        try:
            result = self.client.get_wallet_balance(accountType=ACCOUNT_TYPE, coin="USDT")
            balance = float(result['result']['list'][0]['coin'][0]['walletBalance'])
            return round(balance, 3)
        except Exception as e:
            print(f"get_balance error: {e}")
            return 0.0

    def get_positions(self):
        try:
            result = self.client.get_positions(category="linear", settleCoin="USDT")
            return [pos['symbol'] for pos in result['result']['list']]
        except Exception as e:
            print(f"get_positions error: {e}")
            return []

    def check_closed_pnl(self):
        try:
            df = pd.read_excel('closed_positions_pnl.xlsx')
        except FileNotFoundError:
            df = pd.DataFrame(columns=['Symbol', 'UpdatedTime','Position', 'ClosedPnl'])

        pnl_list = self.client.get_closed_pnl(category="linear", limit=MAX_POSITIONS)['result']['list']

        if pnl_list:
            for position in pnl_list:
                symbol = position['symbol']
                closed_pnl = float(position['closedPnl']) 
                updated_time = position['updatedTime']
                side = position['side'].capitalize()
                position_type = 'LONG' if side == 'Buy' else 'SHORT'

                if int(updated_time) <= TIME_THRESHOLD:
                    continue
                updated_time = datetime.utcfromtimestamp(int(updated_time) / 1000).strftime('%Y-%m-%d %H:%M:%S')

                if df[(df['Symbol'] == symbol) & (df['UpdatedTime'] == updated_time)].empty:
                    new_row = pd.DataFrame({
                        'Symbol': [symbol],
                        'ClosedPnl': [closed_pnl],
                        'UpdatedTime': [updated_time],
                        'Position': [position_type] 
                    })
                    df = pd.concat([df, new_row], ignore_index=True)
                    df.to_excel('closed_positions_pnl.xlsx', index=False)
                    msg = (
                        f"[CLOSED] {symbol} | "
                        f"Closed PnL: {closed_pnl} | "
                        f"Updated Time: {updated_time}"
                    )
                    print(msg)
                    send_discord(msg)
                    self.calculate_total_pnl()
            
            wb = load_workbook('closed_positions_pnl.xlsx')
            ws = wb.active

            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row in range(2, len(df) + 2):
                pnl_cell = ws.cell(row=row, column=4)
                if pnl_cell.value is not None:
                    if pnl_cell.value < 0:
                        pnl_cell.fill = red_fill 
                    elif pnl_cell.value > 0:
                        pnl_cell.fill = green_fill
                    else:
                        pnl_cell.fill = yellow_fill 
            
            wb.save('closed_positions_pnl.xlsx')
        else:
            print("No closed positions found.")
    
    def calculate_total_pnl(self):
        df = pd.read_excel('closed_positions_pnl.xlsx')
        total_pnl = df['ClosedPnl'].sum()
        msg = f"`Total Profit/Loss: {total_pnl} USDT 💵💵💵`"
        print(msg)
        send_discord(msg)

    def get_symbols(self):
        try:
            result = self.client.get_tickers(category="linear")
            return [s['symbol'] for s in result['result']['list'] if "USDT" in s['symbol'] and "USDC" not in s['symbol']]
        except Exception as e:
            print(f"get_symbols error: {e}")
            return []

    def get_mark_price(self, symbol):
        try:
            result = self.client.get_tickers(category="linear", symbol=symbol)
            return float(result['result']['list'][0]['markPrice'])
        except Exception as e:
            print(f"get_mark_price error: {e}")
            return 0.0

    def get_klines(self, symbol, timeframe, limit=500):
        try:
            resp = self.client.get_kline(
                category='linear',
                symbol=symbol,
                interval=timeframe,
                limit=limit,
                recv_window=7000
            )['result']['list']
            
            resp = pd.DataFrame(resp)
            resp.columns = ['timestamp', 'open', 'high', 'low', 'close', 'volume', 'turnover']
            resp = resp.set_index('timestamp')
            resp = resp.astype(float)
            resp = resp[::-1]
            return resp
        except Exception as err:
            print(err)

    def get_precisions(self, symbol):
        try:
            result = self.client.get_instruments_info(category="linear", symbol=symbol)
            data = result['result']['list'][0]
            price_precision = len(data['priceFilter']['tickSize'].split('.')[1]) if '.' in data['priceFilter']['tickSize'] else 0
            qty_precision = len(data['lotSizeFilter']['qtyStep'].split('.')[1]) if '.' in data['lotSizeFilter']['qtyStep'] else 0
            return price_precision, qty_precision
        except Exception as e:
            print(f"get_precisions error: {e}")
            return 0, 0

    def get_max_leverage(self, symbol):
        try:
            result = self.client.get_instruments_info(category="linear", symbol=symbol)
            return float(result['result']['list'][0]['leverageFilter']['maxLeverage'])
        except Exception as e:
            print(f"get_max_leverage error: {e}")
            return 0.0

    def set_mode(self, symbol, mode=1, leverage=10):
        try:
            resp = self.client.switch_margin_mode(category="linear", symbol=symbol, tradeMode=str(mode),
                                                  buyLeverage=str(leverage), sellLeverage=str(leverage))
            if resp['retMsg'] == 'OK':
                mode_str = 'ISOLATED' if mode == 1 else 'CROSS'
                print(f"[{symbol}] Margin mode set to {mode_str}")
        except Exception as e:
            if '110026' in str(e):
                print(f"[{symbol}] Margin mode unchanged")
            else:
                print(f"set_mode error: {e}")

    def set_leverage(self, symbol, leverage=10):
        try:
            resp = self.client.set_leverage(category="linear", symbol=symbol,
                                            buyLeverage=str(leverage), sellLeverage=str(leverage))
            if resp['retMsg'] == 'OK':
                print(f"[{symbol}] Leverage set to {leverage}")
        except Exception as e:
            if '110043' in str(e):
                print(f"[{symbol}] Leverage unchanged")
            else:
                print(f"set_leverage error: {e}")

    def place_market_order(self, symbol, side, mode, leverage, qty, tp, sl, trailing):
        self.set_mode(symbol, mode, leverage)
        sleep(0.5)
        self.set_leverage(symbol, leverage)
        sleep(0.5)

        price_precision, qty_precision = self.get_precisions(symbol)
        order_qty = round(qty, qty_precision)
        tp = round(tp, price_precision) if tp else None
        sl = round(sl, price_precision) if sl else None
        trailing_stop = str(round(float(trailing), price_precision))

        print(f"Placing {side.upper()} market order on {symbol}")
        try:
            resp = self.client.place_order(
                category="linear",
                symbol=symbol,
                side=side.capitalize(),
                orderType="Market",
                qty=order_qty,
                takeProfit=tp,
                stopLoss=sl,
                tpTriggerBy="LastPrice",
                slTriggerBy="LastPrice",
                trailingStop=trailing_stop,    
            )
            print(resp['retMsg'])
        except Exception as e:
            print(f"place_market_order error: {e}")
            
    def set_trailing_stop(self, symbol, trailing_stop):
        try:
            trailing_stop_str = str(trailing_stop)
            resp = self.client.set_trading_stop(
                category="linear",
                symbol=symbol,
                trailingStop=trailing_stop_str
            )
            if resp['retMsg'] == 'OK':
                print(f"Trailing stop updated for {symbol}")
            else:
                print(f"Error setting trailing stop: {resp['retMsg']}")
        except Exception as e:
            print(f"set_trailing_stop error: {e}")
            
    def get_position_details(self, symbol):
        try:
            result = self.client.get_positions(category="linear", settleCoin="USDT")
            
            for pos in result['result']['list']:
                if pos['symbol'] == symbol:
                    return pos 

            print(f"No position found for {symbol}")
            return None
        
        except Exception as e:
            print(f"get_position_details error: {e}")
            return None